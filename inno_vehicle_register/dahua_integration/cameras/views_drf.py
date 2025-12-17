from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.views.generic import TemplateView
from .models import Camera, Vehicle, VehicleInfoByExc
from NetSDK.NetSDK import NetClient
from NetSDK.SDK_Struct import *
from NetSDK.SDK_Enum import *
from NetSDK.SDK_Callback import *
from .sdk_function import *
from .forms import CameraForm
# from .anpr_function import *
from .get_hourly_vehicle import *
from .get_count_vehicle import *
from .capture_process import capture_process
from openpyxl import *
from django.contrib import messages
import sys
from datetime import datetime
from openpyxl.utils import get_column_letter
from rest_framework import generics, status
from .serializers import CameraSerializer, VehicleSerializer, BufferTableSerializer
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
import pandas as pd


callback_num = 0

class CameraListView(generics.ListAPIView):
    queryset = Camera.objects.all()
    serializer_class = CameraSerializer


class BufferTableListView(generics.ListAPIView):
    queryset = BufferTable.objects.all()
    serializer_class = BufferTableSerializer


class VehicleListView(generics.ListAPIView):
    queryset = Vehicle.objects.all().order_by('-timestamp')
    serializer_class = VehicleSerializer


class VehicleTrailerListView(generics.ListCreateAPIView):
    """
    This view handles listing Vehicle objects and creating new ones.
    """
    serializer_class = VehicleSerializer

    def get_queryset(self):
        """
        Override the queryset to filter vehicles where is_trailer_check is True.
        """
        queryset = Vehicle.objects.all().order_by('-timestamp')
        is_trailer_check = self.request.query_params.get('is_trailer_check', None)
        if is_trailer_check is not None and is_trailer_check.lower() == 'true':
            queryset = queryset.filter(is_trailer_check=True)
        return queryset

    def post(self, request, *args, **kwargs):
        """
        Handle POST requests to create or modify vehicle data.
        """
        data = request.data
        is_trailer = data.get('is_trailer', False)

        # Validate and update vehicle_type if necessary
        if is_trailer:
            vehicle_type = data.get('vehicle_type', None)
            if vehicle_type == 4:
                data['vehicle_type'] = 5
            elif vehicle_type == 6:
                data['vehicle_type'] = 7

        serializer = self.serializer_class(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AddBufferTableView(APIView):
    def post(self, request, *args, **kwargs):
        camera_id = request.data.get('camera_id')
        timestamp = request.data.get('timestamp')
        vehicle_color = request.data.get('vehicle_color')
        vehicle_image = request.data.get('vehicle_image')
        plate_image = request.data.get('plate_image')

        try:

            buffer_obj = BufferTable.objects.create(
                camera_id=camera_id, 
                timestamp=timestamp, 
                plate_image=plate_image, 
                vehicle_image=vehicle_image, 
                status= "not processed",
                vehiclecolor=vehicle_color)
            
            print("Object created successfully:", buffer_obj)
            return Response({"message": "Buffer table added successfully."}, status=status.HTTP_200_OK)

        except IntegrityError as e:
            print("Object creation failed:", e)
            return Response({"message": "Buffer table object creation has failed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





class CaptureProcessView(APIView):
    def get(self, request, *args, **kwargs):

        result = capture_process()

        if result:
            return Response({"message": "Buffer table processing completed successfully."}, status=status.HTTP_200_OK)
        else:
            return Response({"message": "Buffer table processing have failed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AlertedVehicleView(generics.ListAPIView):
    queryset = Vehicle.objects.filter(platenumber_fix_req=True)
    serializer_class = VehicleSerializer

class DeleteCameraView(generics.GenericAPIView):
    def post(self, request, pk, *args, **kwargs):
        try:
            print("pk:", pk)
            camera = Camera.objects.get(pk=pk)
            camera.delete()
            return Response({"success": "Camera deleted successfully"}, status=status.HTTP_200_OK)
        except Camera.DoesNotExist:
            return Response({"error": "Camera not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": f"Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# camera connection 
    
class ConnectCameraView(generics.RetrieveAPIView):
    def get(self, request, pk):
        print("connect_camera function ---> START")
        print("pk: ", pk)

        #get camera object
        camera = get_object_or_404(Camera, pk=pk)

        ip = camera.ip_address
        print("ip: ", ip)
        port = camera.port
        print("Port: ", port)
        username = camera.username
        print("username: ", username)
        password = camera.password
        print("password: ", password)

        #SDK function camera login
        loginID, device_info, error_msg, dwAlarmType = camera_login(ip, port, username, password)

        print("loginID: ", loginID)
        print("device_info: ", device_info)
        print("error_msg: ", error_msg)


        if len(error_msg) == 0:
            print("Successful login")
            item = get_object_or_404(Camera, pk=pk)
            print("item_status: ", item.status)
            item.status= 'online'
            item.save()
            message = "Camera connected successfully"
            return Response({'message': message})
        else:
            print("Unsuccessful login")
            item = get_object_or_404(Camera, pk=pk)
            print("item_status: ", item.status)
            item.status = 'offline'
            item.save()
            error_message = "Error connecting to camera"
            return Response({'error': error_message}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
class SubscribeCameraView(APIView):
    def get(self, request, pk):
        print("subscribe_camera function ---> START")

        camera = get_object_or_404(Camera, pk=pk)

        ip = camera.ip_address
        print("ip: ", ip)
        port = camera.port
        print("Port: ", port)
        username = camera.username
        print("username: ", username)
        password = camera.password
        print("password: ", password)

        loginID, device_info, error_msg, dwAlarmType = camera_login(ip, port, username, password)
        camera = get_object_or_404(Camera, pk=pk)
        print("pk: ", pk)
        dwUser = pk
        print("camera: ", camera)
        print('camera_status: ', camera.status)

        if camera.status == "online":
            attachID = camera_subscribe(loginID, device_info, error_msg, dwAlarmType, dwUser)
            print("attachID: ", attachID)

            if attachID > 0:
                camera.subscribe = True
                camera.save()
                message = "Successful subscription!"
                return Response({"message": message})
            else:
                camera.subscribe = False
                camera.save()
                error_msg = "Unsuccessful subscription!"
                return Response({"error": error_msg}, status=status.HTTP_501_NOT_IMPLEMENTED)
        else:
            # Reusing the ConnectCameraView class-based view
            connect_view = ConnectCameraView()
            return connect_view.post(request, pk)


class AddCameraView(generics.CreateAPIView):
    serializer_class = CameraSerializer

    def post(self, request, *args, **kwargs):
        form = CameraForm(request.data)
        print("formdata:", request.data )
        if form.is_valid():
            print("is_111")
            new_camera = form.save(commit=False)
            new_camera.save()
            return Response({"message": "Camera added successfully"}, status= status.HTTP_201_CREATED)
        else:
            print("babab")
            return Response({"error": form.errors}, status=status.HTTP_400_BAD_REQUEST)
        

#class view that will be updating saved 
class UpdateAlertedVehicleView(generics.UpdateAPIView):
    queryset = Vehicle.objects.filter(platenumber_fix_req=True)
    serializer_class = VehicleSerializer
    lookup_url_kwarg = 'pk'

        


def serialize_object(obj):
    # Convert datetime objects to strings with timezone information stripped
    return {
        'timestamp': obj['timestamp'].astimezone().replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S') if 'timestamp' in obj else None,
        'camera_id': obj.get('camera_id'),
        'platenumber': obj.get('platenumber'),
        'vehiclecolor': obj.get('vehiclecolor'),
        'vehicletype': obj.get('vehicletype'),
    }

# class FileUploadView(APIView):
#     parser_classes = [MultiPartParser]

#     def post(self, request, *args, **kwargs):
#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error": "No file uploaded"}, status=400)

#         try:
#             # Read the Excel file using pandas
#             data = pd.read_excel(file)
#             # Perform operations with `data`
#             print(data.head())  # Example: print the first few rows

#             return Response({"message": "File processed successfully!"})
#         except Exception as e:
#             return Response({"error": str(e)}, status=500)


class FileUploadView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=400)

        try:
            # Read the Excel file using pandas
            data = pd.read_excel(file, dtype={'platenumber': str})

            # Rename columns to match model fields
            data.rename(columns={
                "timestamp": "timestamp",
                "location": "location",
                "camera direction": "camera_direction",
                "platenumber": "plate_number"
            }, inplace=True)

            # Convert timestamp to proper format
            data['timestamp'] = pd.to_datetime(data['timestamp'], errors='coerce')

            # Drop rows where timestamp conversion failed
            data = data.dropna(subset=['timestamp'])

            # Convert to model objects
            vehicle_objects = [
                VehicleInfoByExc(
                    timestamp=row['timestamp'],
                    location=row['location'],
                    camera_direction=int(row['camera_direction']),
                    plate_number=row['plate_number']
                ) for _, row in data.iterrows()
            ]

            # Bulk insert
            VehicleInfoByExc.objects.bulk_create(vehicle_objects)

            return Response({"message": "File processed and data saved successfully!"}, status=201)

        except Exception as e:
            return Response({"error": str(e)}, status=500)


class GetReportView(TemplateView):
    def get(self, request, *args, **kwargs):
        if request.method == "GET":
            start_time = request.data.get('start_time')
            end_time = request.data.get('end_time')
            # start_time = "2023-12-31T10:59:59"
            # end_time = "2023-01-01T00:00:00"
            print("test1")
            if start_time is None or end_time is None:
                return Response({"detail": "start_time and end_time are required parameters"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                queryset = get_hourly_vehicle(start_time, end_time)

                # Create the excel file
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=report.xlsx'

                wb = Workbook()
                ws = wb.active
                ws.title = "Report"

                headers = ["timestamp", "camera_id", "platenumber", "vehicle_color", "vehicle_type"]
                ws.append(headers)

                for vehicle in queryset:
                    serialized_vehicle = serialize_object({
                        'timestamp': vehicle.timestamp,
                        'camera_id': vehicle.camera_id,
                        'platenumber': vehicle.platenumber,
                        'vehicle_color': vehicle.vehiclecolor,
                        'vehicle_type': vehicle.vehicletype,
                    })
                    ws.append([serialized_vehicle[key] for key in headers])

                wb.save(response)

                return response
        
            except Exception as e:
                    return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({"detail": "Invalid request method"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
    


class GetDailyReportView(TemplateView):
    def get(self, request, *args, **kwargs):
        try:
            # test = HttpResponse('test', *args, **kwargs)
            # return test
            if request.method == 'GET':
                # Get the selected date from the form submission
                selected_date_str = request.data.get('selected_date')
                selected_camera = request.data.get('selected_camera')
                # selected_date_str = '2024-02-13'
                # selected_camera = 17

                # Result is given within List with dict. timerange as keys and values are count
                result_list = get_daily_count_list(selected_date_str, selected_camera)


                # Load the Excel template
                template_path = 'C:\\Users\\bilgee work\\Desktop\\dahua sdk django\\dahua_integration\\template.xlsx'
                template_wb = load_workbook(template_path)
                template_wb.properties.calculate_formula = True
                template_ws = template_wb.active

                print('result_list: ', result_list)

                

                # Update sheets
                for time_range in range(1, 25):
                    for timerange_dict in result_list:
                        counts = timerange_dict.get(str(time_range), [])
                        start_column = 'C'
                        start_row = get_start_row(time_range)

                        # Update cells with the values from the list
                        for index, value in enumerate(counts):
                            column_letter = chr(ord(start_column) + index)
                            cell_coordinates = f"{column_letter}{start_row}"
                            template_ws[cell_coordinates] = value

                # Update the date in excel
                date_cell = 'M6'
                template_ws[date_cell] = selected_date_str

                # Calculate sum of the counts
                start_row = 14
                end_row = 37

                # Loop through each column and add SUM formula in row 38
                for col_idx in range(3, 14):
                    col_letter = get_column_letter(col_idx)
                    sum_formula = f'=SUM({col_letter}{start_row}:{col_letter}{end_row})'
                    template_ws[f'{col_letter}38'] = sum_formula

                # Add the sum of column C to M and store in column N for row 38
                template_ws['N38'] = f'=SUM(C38:M38)'

                template_wb.properties.calculate_formula = True

                # Save the new workbook
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=daily_report.xlsx'
                template_wb.save(response)

                # Return the new workbook as a response
                return response

        except Exception as e:
            # Handle exceptions and return an error response
            return Response({"error": f"Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GetDailyReportAllCamView(TemplateView):
    def get(self, request, *args, **kwargs):
        try:
            # test = HttpResponse('test', *args, **kwargs)
            # return test
            if request.method == 'GET':
                # Get the selected date from the form submission
                selected_date_str = request.data.get('selected_date')
                # selected_camera = request.data.get('selected_camera')

                # Get all the cameras from the DB
                # iterate through each cameras and get the counts
                cameras_all = Camera.objects.all()
                cam_ids = []
                for camera in cameras_all:
                    cam_id = camera['id']
                    cam_ids.append(cam_id)

                # selected_date_str = '2024-02-13'
                for cam_id in cam_ids:
                    selected_camera = cam_id
                    # selected_camera = 17

                    # Result is given within List of dict. list elements start from type 1 vehicle to type 9 vehicle.
                    result_list = get_daily_count_list(selected_date_str, selected_camera)

                    print('result_list: ', result_list)

                    #Since the Result is given from with list of dict. [{1:5, 2:6, 3:5, 4:0}; {1:5, 2:6, 3:5, 4:0}; {1:5, 2:6, 3:5, 4:0}]
                    #Put the result into a another dictionary. key as camera id and value as count.
                    cam_daily_count_dict = {str(selected_camera) : result_list}
                    cam_daily_count_list = []
                    cam_daily_count_list.append(cam_daily_count_dict)

                    
                # Response should be including all the camera counts
                # Return the cam_daily_count_list
                return cam_daily_count_list



        except Exception as e:
            # Handle exceptions and return an error response
            return Response({"error": f"Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
# class GetDashboardInfoView(APIView):
#     def get()


class PostVehicleInfoView(APIView):
    def post(self, request, *args, **kwargs):
        # The request should be from ТАЦ system providers
        try:
            # Get information from the request
            camera_id = request.data.get('camera_id')
            timestamp = request.data.get('timestamp')
            vehiclecolor = request.data.get('vehiclecolor')
            vehicle_image = request.data.get('vehicle_image')
            platenumber = request.data.get('platenumber')
            vehicletype = request.data.get('vehicletype')
            plate_image = request.data.get('plate_image')


            # create vehicle object 
            vehicle_obj = Vehicle.objects.create(
                camera_id=camera_id,
                timestamp=timestamp,
                vehiclecolor=vehiclecolor,
                vehicle_image=vehicle_image,
                platenumber=platenumber,
                vehicletype=vehicletype,
                plate_image=plate_image,
            )
            # On successful creation of vehicle. return response message.
            return Response({"message": "Vehicle created successfully."}, status=status.HTTP_201_CREATED)

        except IntegrityError as e:
            return Response({"error": "Vehicle creation failed.", "details": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            return Response({"error": "An error occurred.", "details": str(e)}, status=status.HTTP_400_BAD_REQUEST)